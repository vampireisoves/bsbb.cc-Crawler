"""
BSBB节点爬虫 - 从bsbb.cc网站爬取V2Ray节点信息
功能：
- 下载节点数据
- 解析节点信息（国家、协议、主机、名称、URL）
- 保存到Excel文件
- 生成README.md统计报告
"""

import requests
import re
import base64
import json
from urllib.parse import unquote
from datetime import datetime, timedelta, timezone

# 国家映射
COUNTRY_MAP = {
    'us': 'United States',
    'hk': 'Hong Kong',
    'jp': 'Japan',
    'sg': 'Singapore',
    'kr': 'South Korea',
    'de': 'Germany',
    'fr': 'France',
    'gb': 'United Kingdom',
    'ca': 'Canada',
    'au': 'Australia',
    'nl': 'Netherlands',
    'it': 'Italy',
    'ru': 'Russia',
    'tr': 'Turkey',
    'br': 'Brazil',
    'in': 'India',
    'id': 'Indonesia',
    'my': 'Malaysia',
    'th': 'Thailand',
    'vn': 'Vietnam',
    'ph': 'Philippines',
    'es': 'Spain',
    'se': 'Sweden',
    'ch': 'Switzerland',
    'pl': 'Poland',
    'ua': 'Ukraine',
    'mx': 'Mexico',
    'ae': 'United Arab Emirates'
}

FLAG_TO_COUNTRY = {
    '🇺🇸': 'United States',
    '🇭🇰': 'Hong Kong',
    '🇯🇵': 'Japan',
    '🇸🇬': 'Singapore',
    '🇰🇷': 'South Korea',
    '🇩🇪': 'Germany',
    '🇫🇷': 'France',
    '🇬🇧': 'United Kingdom',
    '🇨🇦': 'Canada',
    '🇦🇺': 'Australia',
    '🇳🇱': 'Netherlands',
    '🇮🇹': 'Italy',
    '🇷🇺': 'Russia',
    '🇹🇷': 'Turkey',
    '🇧🇷': 'Brazil',
    '🇮🇳': 'India',
    '🇮🇩': 'Indonesia',
    '🇲🇾': 'Malaysia',
    '🇹🇭': 'Thailand',
    '🇻🇳': 'Vietnam',
    '🇵🇭': 'Philippines',
    '🇪🇸': 'Spain',
    '🇸🇪': 'Sweden',
    '🇨🇭': 'Switzerland',
    '🇵🇱': 'Poland',
    '🇺🇦': 'Ukraine',
    '🇲🇽': 'Mexico',
    '🇦🇪': 'United Arab Emirates'
}

# 国家英文名到中文名的映射（用于README生成）
COUNTRY_NAME_MAP = {
    'United States': '美国',
    'Hong Kong': '香港',
    'Japan': '日本',
    'Singapore': '新加坡',
    'South Korea': '韩国',
    'Germany': '德国',
    'France': '法国',
    'United Kingdom': '英国',
    'Canada': '加拿大',
    'Australia': '澳大利亚',
    'Netherlands': '荷兰',
    'Italy': '意大利',
    'Russia': '俄罗斯',
    'Turkey': '土耳其',
    'Brazil': '巴西',
    'India': '印度',
    'Indonesia': '印尼',
    'Malaysia': '马来西亚',
    'Thailand': '泰国',
    'Vietnam': '越南',
    'Philippines': '菲律宾',
    'Spain': '西班牙',
    'Sweden': '瑞典',
    'Switzerland': '瑞士',
    'Poland': '波兰',
    'Ukraine': '乌克兰',
    'Mexico': '墨西哥',
    'United Arab Emirates': '阿联酋',
    'Taiwan': '台湾',
    'Unknown': '未知'
}


def get_first_flag_emoji(text):
    """从文本中提取第一个国旗emoji"""
    cps = list(text)
    for i in range(len(cps) - 1):
        a = ord(cps[i])
        b = ord(cps[i + 1])
        if 0x1F1E6 <= a <= 0x1F1FF and 0x1F1E6 <= b <= 0x1F1FF:
            return cps[i] + cps[i + 1]
    return None


def extract_country_from_name(name):
    """从节点名称中提取中文国家名"""
    if not name:
        return None
    
    # 中文国家名映射
    cn_country_map = {
        '香港': 'Hong Kong',
        '日本': 'Japan',
        '韓國': 'South Korea',
        '韩国': 'South Korea',
        '法國': 'France',
        '法国': 'France',
        '美國': 'United States',
        '美国': 'United States',
        '新加坡': 'Singapore',
        '台湾': 'Taiwan',
        '臺灣': 'Taiwan',
        '英國': 'United Kingdom',
        '英国': 'United Kingdom',
        '德國': 'Germany',
        '德国': 'Germany',
        '加拿大': 'Canada',
        '澳洲': 'Australia',
        '澳大利亚': 'Australia',
        '俄罗斯': 'Russia',
        '俄羅斯': 'Russia',
        '土耳其': 'Turkey',
        '巴西': 'Brazil',
        '印度': 'India',
        '印尼': 'Indonesia',
        '馬來西亞': 'Malaysia',
        '马来西亚': 'Malaysia',
        '泰國': 'Thailand',
        '泰国': 'Thailand',
        '越南': 'Vietnam',
        '菲律宾': 'Philippines',
        '西班牙': 'Spain',
        '瑞典': 'Sweden',
        '瑞士': 'Switzerland',
        '波兰': 'Poland',
        '乌克兰': 'Ukraine',
        '墨西哥': 'Mexico',
        '阿联酋': 'United Arab Emirates'
    }
    
    # 检查是否包含国家名
    for cn_name, en_name in cn_country_map.items():
        if cn_name in name:
            return en_name
    
    return None


def get_protocol(line):
    """获取协议类型"""
    lower_line = line.lower()
    idx = lower_line.find('://')
    if idx < 0:
        return None
    start = lower_line.rfind(' ', idx - 1) + 1
    proto = lower_line[start:idx]
    valid_protocols = ['vless', 'vmess', 'trojan', 'ss', 'ssr', 'hysteria', 'hysteria2']
    return proto if proto in valid_protocols else None


def get_url(line, proto):
    """获取URL"""
    token = proto + '://'
    lower_line = line.lower()
    i = lower_line.find(token)
    if i < 0:
        return None
    j = line.find(' ', i)
    if j < 0:
        j = len(line)
    return line[i:j].strip()


def parse_host_port_sni(url, proto):
    """解析host、port和sni"""
    host = None
    port = None
    sni = None
    
    if not url:
        return {'host': host, 'port': port, 'sni': sni}
    
    # vmess协议特殊处理
    if proto == 'vmess':
        try:
            b64 = url[8:]
            b64 = re.sub(r'\s+', '', b64)
            b64 = b64.replace('_', '/').replace('-', '+')
            pad = b64 + '=' * ((4 - len(b64) % 4) % 4)
            obj = json.loads(base64.b64decode(pad))
            host = obj.get('add') or obj.get('host')
            port = obj.get('port')
        except:
            pass
        return {'host': host, 'port': port, 'sni': sni}
    
    # 其他协议处理
    r = url
    idx = r.find('://')
    if idx > -1:
        r = r[idx + 3:]
    
    at = r.find('@')
    if at > -1:
        r = r[at + 1:]
    
    cut = len(r)
    pI = r.find('/')
    qI = r.find('?')
    
    if pI > -1:
        cut = min(cut, pI)
    if qI > -1:
        cut = min(cut, qI)
    
    hp = r[:cut]
    c = hp.find(':')
    
    if c > -1:
        host = hp[:c]
        port = hp[c + 1:]
    else:
        host = hp
    
    if qI > -1:
        params = r[qI + 1:].split('&')
        for kv in params:
            parts = kv.split('=')
            if len(parts) == 2:
                k, v = parts
                val = unquote(v)
                if not sni and k in ['sni', 'serverName', 'peer']:
                    sni = val
    
    return {'host': host, 'port': port, 'sni': sni}


def get_latency_ms(line):
    """获取延迟时间"""
    lower_line = line.lower()
    i = lower_line.find('ms')
    if i < 0:
        return None
    k = i - 1
    d = ''
    while k >= 0:
        c = lower_line[k]
        if '0' <= c <= '9':
            d = c + d
            k -= 1
            continue
        if c == ' ':
            k -= 1
            continue
        break
    try:
        return int(d) if d else None
    except:
        return None


def download_nodes():
    """下载节点数据"""
    urls = [
        'https://www.bsbb.cc/V2RAY.txt',
        'https://www.bsbb.cc/V22RAY.txt'
    ]
    
    all_lines = []
    
    for url in urls:
        print(f"正在下载: {url}")
        try:
            response = requests.get(url, timeout=30)
            if response.status_code == 200:
                lines = response.text.strip().split('\n')
                all_lines.extend([line.strip() for line in lines if line.strip()])
                print(f"  下载成功，获取 {len(lines)} 行")
            else:
                print(f"  下载失败，状态码: {response.status_code}")
        except Exception as e:
            print(f"  下载错误: {e}")
    
    return all_lines


def parse_nodes(lines):
    """解析节点数据"""
    nodes = []
    
    for line in lines:
        # 提取协议
        proto = get_protocol(line)
        if not proto:
            continue
        
        # 提取URL
        url = get_url(line, proto)
        if not url:
            continue
        
        # 提取延迟
        latency = get_latency_ms(line)
        
        # 提取名称
        name = None
        if '#' in url:
            name = unquote(url.split('#')[-1])
        
        # 提取国家信息（优先级：名称中的中文 > URL中的国旗 > 原始行中的国旗）
        country = 'Unknown'
        
        # 1. 从名称中提取中文国家名
        if name:
            name_country = extract_country_from_name(name)
            if name_country:
                country = name_country
        
        # 2. 如果没有找到，从URL中提取国旗
        if country == 'Unknown' and url:
            flag = get_first_flag_emoji(url)
            if flag and flag in FLAG_TO_COUNTRY:
                country = FLAG_TO_COUNTRY[flag]
        
        # 3. 如果还是没有找到，从原始行中提取国旗
        if country == 'Unknown':
            flag = get_first_flag_emoji(line)
            if flag and flag in FLAG_TO_COUNTRY:
                country = FLAG_TO_COUNTRY[flag]
        
        # 解析host、port、sni
        host_info = parse_host_port_sni(url, proto)
        
        node = {
            'Country': country,
            'Protocol': proto.upper(),
            'Host': host_info['host'] if host_info['host'] else 'N/A',
            'Name': name if name else 'N/A',
            'URL': url,
            'Latency': f"{latency}ms" if latency else 'N/A'
        }
        
        nodes.append(node)
    
    return nodes


def save_to_txt(nodes, filename):
    """保存到文本文件"""
    with open(filename, 'w', encoding='utf-8') as f:
        f.write("Country\tProtocol\tHost\tName\tURL\n")
        f.write("-" * 100 + "\n")
        for node in nodes:
            f.write(f"{node['Country']}\t{node['Protocol']}\t{node['Host']}\t{node['Name']}\t{node['URL']}\n")
    
    print(f"\n已保存 {len(nodes)} 个节点到 {filename}")


def save_to_xlsx(nodes, filename):
    """保存到Excel文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "节点信息"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 80
    
    # 设置标题样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题
    headers = ['Country', 'Protocol', 'Host', 'Name', 'URL']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置数据样式
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 写入数据
    for row, node in enumerate(nodes, 2):
        ws.cell(row=row, column=1, value=node['Country']).font = data_font
        ws.cell(row=row, column=1).alignment = data_alignment
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value=node['Protocol']).font = data_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3, value=node['Host']).font = data_font
        ws.cell(row=row, column=3).alignment = data_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4, value=node['Name']).font = data_font
        ws.cell(row=row, column=4).alignment = data_alignment
        ws.cell(row=row, column=4).border = thin_border
        
        ws.cell(row=row, column=5, value=node['URL']).font = data_font
        ws.cell(row=row, column=5).alignment = data_alignment
        ws.cell(row=row, column=5).border = thin_border
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    wb.save(filename)
    print(f"\n已保存 {len(nodes)} 个节点到 {filename}")


def save_to_b64_config(nodes, filename):
    """提取所有URL，进行base64编码后保存为纯文本文件"""
    # 提取所有URL
    urls = [node['URL'] for node in nodes if node['URL']]
    
    # 用换行符连接所有URL
    urls_text = '\n'.join(urls)
    
    # 进行base64编码
    b64_encoded = base64.b64encode(urls_text.encode('utf-8')).decode('utf-8')
    
    # 只保存base64编码后的字符串到文件
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(b64_encoded)
    
    print(f"\n已保存 {len(urls)} 个节点URL到 {filename}")
    print(f"  Base64编码长度: {len(b64_encoded)} 字符")


def get_country_name_cn(english_name):
    """获取国家中文名称"""
    return COUNTRY_NAME_MAP.get(english_name, english_name)


def generate_readme():
    """生成README.md统计报告"""
    from openpyxl import load_workbook
    
    print("\n正在生成README.md统计报告...")
    
    # 读取Excel文件
    wb = load_workbook('node.xlsx')
    ws = wb.active
    
    nodes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # 只读取有数据的行
            nodes.append({
                'Country': row[0],
                'Protocol': row[1],
                'Host': row[2],
                'Name': row[3],
                'URL': row[4]
            })
    
    # 生成统计
    country_stats = {}
    for node in nodes:
        country = node['Country']
        country_cn = get_country_name_cn(country)
        country_stats[country_cn] = country_stats.get(country_cn, 0) + 1
    
    protocol_stats = {}
    for node in nodes:
        protocol = node['Protocol']
        protocol_stats[protocol] = protocol_stats.get(protocol, 0) + 1
    
    # 获取北京时间（UTC+8）
    beijing_tz = timezone(timedelta(hours=8))
    crawl_time = datetime.now(beijing_tz).strftime('%Y-%m-%d %H:%M:%S')
    
    # 生成README内容
    content = f"""# BSBB节点爬虫统计报告

## 爬取信息

- **数据源**: https://bsbb.cc/
- **爬取时间**: {crawl_time} (北京时间)
- **时区**: UTC+8

## 爬虫结果统计

- **总节点数**: {len(nodes)} 个
- **国家/地区数量**: {len(country_stats)} 个
- **协议类型数量**: {len(protocol_stats)} 种

## 按国家/区域统计

| 国家/地区 | 节点数量 | 占比 |
|---------|---------|------|
"""
    
    # 按节点数量降序排列
    sorted_countries = sorted(country_stats.items(), key=lambda x: x[1], reverse=True)
    for country, count in sorted_countries:
        percentage = (count / len(nodes)) * 100
        content += f"| {country} | {count} | {percentage:.1f}% |\n"
    
    content += f"""
## 按协议类型统计

| 协议类型 | 节点数量 | 占比 |
|---------|---------|------|
"""
    
    # 按节点数量降序排列
    sorted_protocols = sorted(protocol_stats.items(), key=lambda x: x[1], reverse=True)
    for protocol, count in sorted_protocols:
        percentage = (count / len(nodes)) * 100
        content += f"| {protocol} | {count} | {percentage:.1f}% |\n"
    
    content += f"""
## 订阅链接

https://github.com/vampireisoves/BSBBScrape/blob/main/b64config.json

---

*最后更新: {crawl_time}*
"""
    
    # 保存README.md
    with open('README.md', 'w', encoding='utf-8') as f:
        f.write(content)
    
    print("README.md 生成成功！")
    print(f"  总节点数: {len(nodes)}")
    print(f"  国家/地区数量: {len(country_stats)}")
    print(f"  协议类型数量: {len(protocol_stats)}")


def main():
    """主函数"""
    print("=" * 60)
    print("BSBB节点爬虫 v1.0")
    print("从 https://bsbb.cc/ 爬取V2Ray节点信息")
    print("=" * 60)
    print()
    
    print("开始下载节点数据...")
    lines = download_nodes()
    
    print(f"\n总共下载了 {len(lines)} 行数据")
    print("开始解析节点...")
    
    nodes = parse_nodes(lines)
    
    print(f"成功解析 {len(nodes)} 个节点")
    
    save_to_xlsx(nodes, 'node.xlsx')
    
    # 保存base64编码的配置
    save_to_b64_config(nodes, 'b64config.json')
    
    # 统计国家分布
    countries = {}
    for node in nodes:
        c = node['Country']
        countries[c] = countries.get(c, 0) + 1
    
    print("\n" + "=" * 60)
    print("国家分布统计:")
    print("=" * 60)
    for country, count in sorted(countries.items()):
        print(f"  {country}: {count} 个节点")
    
    # 显示前5个节点作为示例
    print("\n" + "=" * 60)
    print("前5个节点示例:")
    print("=" * 60)
    for i, node in enumerate(nodes[:5], 1):
        print(f"\n节点 {i}:")
        print(f"  Country: {node['Country']}")
        print(f"  Protocol: {node['Protocol']}")
        print(f"  Host: {node['Host']}")
        print(f"  Name: {node['Name']}")
        print(f"  URL: {node['URL'][:80]}...")
    
# 生成README.md
    generate_readme()
    
    # 删除 node.xlsx 文件
    import os
    if os.path.exists('node.xlsx'):
        os.remove('node.xlsx')
        print("\n已删除 node.xlsx 文件")
    
    print("\n" + "=" * 60)
    print("爬取完成！结果已保存到:")
    print("  - b64config.json (Base64编码的节点配置)")
    print("  - README.md (统计报告)")
    print("=" * 60)


if __name__ == "__main__":
    main()